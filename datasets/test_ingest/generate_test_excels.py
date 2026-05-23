"""
Genera dos Excel de prueba para el flujo de ingesta IA de SmartSupply.

1. ventas_distribuidora_desordenado.xlsx
   - Columnas con nombres chilenos no-estándar  → dispara path Claude Sonnet
   - Tiene fila de título, celdas combinadas estilo "cuaderno de distribuidora"
   - Algunas filas con datos faltantes para generar warnings

2. ventas_distribuidora_limpio.xlsx
   - Columnas reconocibles directamente por pandas (_try_direct_mapping)
   - Testea el path rápido sin gastar tokens de Claude

Uso: python datasets/test_ingest/generate_test_excels.py
Salida: datasets/test_ingest/*.xlsx
"""

from pathlib import Path
from datetime import date, timedelta
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).parent
random.seed(42)

FAMILIAS = [
    "Lacteos", "Bebidas", "Abarrotes", "Carnes y Embutidos",
    "Limpieza del Hogar", "Panaderia y Pasteleria", "Congelados",
    "Higiene Personal", "Frutas y Verduras",
]

TIENDA = "Distribuidora Santa Elena - Sucursal Ñuñoa"

def fechas(inicio: date, dias: int):
    return [inicio + timedelta(days=i) for i in range(dias)]

def venta_base(familia: str) -> float:
    bases = {
        "Lacteos": 280_000, "Bebidas": 420_000, "Abarrotes": 650_000,
        "Carnes y Embutidos": 380_000, "Limpieza del Hogar": 190_000,
        "Panaderia y Pasteleria": 210_000, "Congelados": 150_000,
        "Higiene Personal": 170_000, "Frutas y Verduras": 320_000,
    }
    base = bases.get(familia, 200_000)
    factor = 0.75 + random.random() * 0.50   # ±25%
    return round(base * factor, 0)


# ─────────────────────────────────────────────────────────────────────────────
# Excel 1 — DESORDENADO (dispara Claude)
# Simula un Excel exportado de un sistema legacy chileno:
#   - Fila 1: título con nombre de la tienda (celdas combinadas)
#   - Fila 2: período del reporte
#   - Fila 3: vacía
#   - Fila 4: encabezados con nombres no-standard
#   - Fila 5+: datos, con columna "Monto Neto ($)" y "Rubro Comercial"
# ─────────────────────────────────────────────────────────────────────────────
def generar_desordenado():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas Mayo 2026"

    DIAS = fechas(date(2025, 11, 24), 180)   # nov 2025 → may 2026 (180 días)
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    TITLE_FONT  = Font(bold=True, size=13, color="1E3A5F")
    ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fila 1 — título
    ws.merge_cells("A1:F1")
    ws["A1"] = TIENDA
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Fila 2 — periodo
    ws.merge_cells("A2:F2")
    ws["A2"] = "Reporte de Ventas — Nov 2025 – May 2026 (al 20/05/2026)"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="555555")

    # Fila 3 — vacía
    ws.row_dimensions[3].height = 6

    # Fila 4 — encabezados ambiguos
    headers = [
        "N° Registro",      # col A  ← ruido
        "Fecha Operacion",  # col B  ← fecha (no-standard)
        "Rubro Comercial",  # col C  ← familia (no-standard)
        "Monto Neto ($)",   # col D  ← ventas (no-standard)
        "Con Oferta",       # col E  ← onpromotion (no-standard)
        "Obs.",             # col F  ← ruido (observaciones vacías)
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20

    # Datos
    registros = []
    for i, dia in enumerate(DIAS):
        # Sábados con volumen diferente, sin domingos (distribuidora cierra)
        if dia.weekday() == 6:  # domingo
            continue
        for familia in FAMILIAS:
            monto = venta_base(familia)
            # ~30% de días con alguna promo
            con_oferta = "Sí" if random.random() < 0.30 else "No"
            registros.append((dia, familia, monto, con_oferta))

    random.shuffle(registros)  # simula que no están ordenados

    for idx, (dia, familia, monto, oferta) in enumerate(registros, start=1):
        row = idx + 4  # datos desde fila 5
        fill = ALT_FILL if idx % 2 == 0 else None

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=dia.strftime("%d/%m/%Y"))   # formato chileno
        ws.cell(row=row, column=3, value=familia)
        ws.cell(row=row, column=4, value=monto)
        ws.cell(row=row, column=4).number_format = "#,##0"
        ws.cell(row=row, column=5, value=oferta)
        ws.cell(row=row, column=6, value="")  # obs vacía

        for col_idx in range(1, 7):
            c = ws.cell(row=row, column=col_idx)
            c.border = border
            if fill:
                c.fill = fill

    # Fila de total al final (ruido extra)
    total_row = len(registros) + 5
    ws.cell(row=total_row, column=3, value="TOTAL PERÍODO").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(r[2] for r in registros))
    ws.cell(row=total_row, column=4).number_format = "#,##0"
    ws.cell(row=total_row, column=4).font = Font(bold=True)

    out = OUTPUT_DIR / "ventas_180dias_desordenado.xlsx"
    wb.save(out)
    print(f"✓ Generado: {out.name}  ({len(registros)} registros, path Claude)")


# ─────────────────────────────────────────────────────────────────────────────
# Excel 2 — LIMPIO (mapeo directo por pandas, sin Claude)
# Columnas: fecha | familia | ventas | onpromotion
# ─────────────────────────────────────────────────────────────────────────────
def generar_limpio():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    DIAS = fechas(date(2025, 11, 1), 180)   # nov 2025 → abr 2026

    headers = ["fecha", "familia", "ventas", "onpromotion"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True)

    row = 2
    for dia in DIAS:
        for familia in FAMILIAS:
            monto = venta_base(familia)
            promo = 1 if random.random() < 0.25 else 0
            ws.cell(row=row, column=1, value=dia.isoformat())
            ws.cell(row=row, column=2, value=familia.upper())
            ws.cell(row=row, column=3, value=monto)
            ws.cell(row=row, column=4, value=promo)
            row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    out = OUTPUT_DIR / "ventas_180dias_limpio.xlsx"
    wb.save(out)
    print(f"✓ Generado: {out.name}  ({(row-2)} registros, path pandas directo)")


if __name__ == "__main__":
    generar_desordenado()
    generar_limpio()
    print("\nArchivos listos en datasets/test_ingest/")
